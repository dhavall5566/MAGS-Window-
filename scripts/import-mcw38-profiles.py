#!/usr/bin/env python3
"""Import MCW38 CURTAIN WALL.xlsx into profiles + series_names tables (with images)."""

from __future__ import annotations

import asyncio
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from datetime import date
from pathlib import Path

import openpyxl

BACKEND_ROOT = Path(__file__).resolve().parents[2] / "Api-MAGS-devansh-main"
sys.path.insert(0, str(BACKEND_ROOT))

from app.database import SessionLocal  # noqa: E402
from app.models import Profile, SeriesName  # noqa: E402
from app.services.db_helpers import upsert_entity  # noqa: E402
from app.services.imagekit import upload_image_bytes  # noqa: E402

RMM_FACTOR = 305
R_MTR_RATE_MULTIPLIER = 3.25
DEFAULT_XLSX = Path.home() / "Downloads" / "MCW38 CURATIN WALL.xlsx"
PROFILE_IMAGE_COL = 4
DRAWING_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}
R_EMBED = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"


def parse_series(series_text: str) -> tuple[str, str, str, str]:
    text = (series_text or "").strip()
    match = re.match(r"^([A-Za-z]+)(\d+)\s*(.*)$", text)
    if match:
        name, series_no, suffix = match.group(1).upper(), match.group(2), match.group(3).strip()
        base = f"{name}{series_no}"
        full = f"{base} {suffix}".strip() if suffix else base
        return name, series_no, suffix, full
    return text[:2].upper(), "", "", text


def extract_row_images(xlsx_path: Path) -> dict[int, bytes]:
    """Map Excel row index (1-based data rows) to embedded PNG bytes in PROFILE IMAGE column."""
    with zipfile.ZipFile(xlsx_path) as z:
        drawing = ET.fromstring(z.read("xl/drawings/drawing1.xml"))
        rels_root = ET.fromstring(z.read("xl/drawings/_rels/drawing1.xml.rels"))
        rel_map = {
            rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/")
            for rel in rels_root.findall("rel:Relationship", DRAWING_NS)
        }

        row_paths: dict[int, str] = {}
        for anchor in drawing.findall("xdr:twoCellAnchor", DRAWING_NS):
            from_cell = anchor.find("xdr:from", DRAWING_NS)
            if from_cell is None:
                continue
            row_el = from_cell.find("xdr:row", DRAWING_NS)
            col_el = from_cell.find("xdr:col", DRAWING_NS)
            if row_el is None or col_el is None:
                continue
            if int(col_el.text) != PROFILE_IMAGE_COL:
                continue
            blip = anchor.find(".//a:blip", DRAWING_NS)
            if blip is None:
                continue
            embed = blip.attrib.get(R_EMBED)
            if embed and embed in rel_map:
                row_paths[int(row_el.text)] = rel_map[embed]

        return {row: z.read(path) for row, path in row_paths.items()}


async def upload_row_images(row_bytes: dict[int, bytes]) -> dict[int, str]:
    urls: dict[int, str] = {}
    for row, content in sorted(row_bytes.items()):
        result = await upload_image_bytes(
            filename=f"mcw38-profile-{str(row).zfill(2)}.png",
            content=content,
            content_type="image/png",
            folder="/mags/profiles/mcw38",
        )
        url = result.get("url")
        if not url:
            raise RuntimeError(f"ImageKit returned no URL for row {row}")
        urls[row] = url
    return urls


def r_mtr_rate(rmm: float, rate: float) -> float:
    if not rmm or not rate:
        return 0.0
    return round(((rmm / RMM_FACTOR) * rate * R_MTR_RATE_MULTIPLIER) * 100) / 100


def build_profile(row: dict, today: str, image_url: str = "") -> dict:
    profile_code = str(row["profile_code"]).strip()
    profile_no = profile_code.split("/")[-1] if "/" in profile_code else "1"
    kg_mtr = round(float(row["kg_mtr"]), 4)
    periferi = round(float(row["periferi"]))
    length_m = round(periferi / RMM_FACTOR, 4) if periferi else 1.0
    if length_m <= 0:
        length_m = 1.0
    profile_name = str(row["profile_name"]).strip()
    rate_per_meter = r_mtr_rate(float(periferi), kg_mtr)

    return {
        "id": f"prf-mcw-{profile_no.zfill(2)}",
        "code": profile_code,
        "name": profile_name,
        "seriesName": row["series_label"],
        "profileNo": profile_no,
        "dyeCode": str(row["dia_code"]).strip(),
        "rmm": length_m,
        "powderCoatingRmm": periferi,
        "ratePerMeter": 0,
        "category": "",
        "alloy": "",
        "finish": "",
        "weightPerMeter": kg_mtr,
        "standardLength": length_m,
        "image": image_url,
        "design": image_url,
        "designName": profile_name,
        "purchaseUnitQty": periferi,
        "purchaseUnitMetric": "KG",
        "conversionUnitQty": kg_mtr,
        "conversionUnitMetric": "Meter",
        "description": profile_name,
        "minStock": 0,
        "currentStock": 0,
        "unit": "pcs",
        "status": "active",
        "createdAt": today,
    }


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [
        str(cell.value).strip().upper() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    idx = {header: i for i, header in enumerate(headers)}

    rows: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or excel_row[idx.get("SR.NO", 0)] in (None, ""):
            continue
        series_text = excel_row[idx["SERIS NAME"]]
        _, _, _, series_label = parse_series(str(series_text))
        rows.append(
            {
                "excel_row": len(rows) + 1,
                "dia_code": excel_row[idx["DIA CODE"]],
                "series_text": series_text,
                "series_label": series_label,
                "profile_code": excel_row[idx["PROFILE CODE"]],
                "profile_name": excel_row[idx["PROFILE NAME"]],
                "kg_mtr": excel_row[idx["KG/MTR"]],
                "periferi": excel_row[idx["PERIFERI"]],
            }
        )
    return rows


async def run_import(xlsx_path: Path) -> None:
    rows = load_rows(xlsx_path)
    if not rows:
        raise SystemExit("No profile rows found in spreadsheet")

    row_images = extract_row_images(xlsx_path)
    image_urls = await upload_row_images(row_images) if row_images else {}

    today = date.today().isoformat()
    series_name, series_no, series_suffix, series_label = parse_series(str(rows[0]["series_text"]))
    series_record = {
        "id": "ser-mcw38",
        "name": series_name,
        "seriesNo": series_no,
        "seriesSuffix": series_suffix,
        "status": "active",
        "createdAt": today,
    }
    profiles = [
        build_profile(row, today, image_urls.get(row["excel_row"], ""))
        for row in rows
    ]

    db = SessionLocal()
    try:
        upsert_entity(db, SeriesName, series_record)
        for profile in profiles:
            upsert_entity(db, Profile, profile)
    finally:
        db.close()

    print(
        f"Imported series {series_name}{series_no}, {len(profiles)} profiles, "
        f"{len(image_urls)} images from {xlsx_path.name}"
    )
    for profile in profiles:
        has_image = "yes" if profile["image"] else "no"
        print(f"  - {profile['code']} image={has_image}")


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        raise SystemExit(f"Excel file not found: {xlsx_path}")
    asyncio.run(run_import(xlsx_path))


if __name__ == "__main__":
    main()
